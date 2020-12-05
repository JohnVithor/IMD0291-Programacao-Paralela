# -*- coding: utf-8 -*-
import pandas as pd
import glob
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

path_serials = list(glob.glob("./*erial*.csv"))
path_parallels = list(glob.glob("./*allel*.csv"))

def remove_min_max(data):
    mins = list(data.groupby(["program", "problem_size"]).idxmin()["execution_time"].values)
    maxs = list(data.groupby(["program", "problem_size"]).idxmax()["execution_time"].values)
    return data.drop(mins+maxs)

def save_with_speedup_efic(dataframe, name):
    workbook = Workbook()
    ws = workbook.active
    dataframe.index
    rows = dataframe_to_rows(dataframe, index=True, header=True)
    ws.append(next(rows))
    next(rows)
    for r in rows:
        ws.append(r)
    ws["A1"] = "Tamanho do Problema"
    start_column = len(dataframe.columns) + 2
    ws.cell(row=1, column=start_column).value = "Speedup 4"
    ws.cell(row=1, column=start_column+1).value = "Speedup 8"
    ws.cell(row=1, column=start_column+2).value = "Speedup 16"
    ws.cell(row=1, column=start_column+3).value = "Speedup 32"

    for i in range(2, 10):
        ws["G"+str(i)] = "=B"+str(i)+" / C"+str(i)
        ws["H"+str(i)] = "=B"+str(i)+" / D"+str(i)
        ws["I"+str(i)] = "=B"+str(i)+" / E"+str(i)
        ws["J"+str(i)] = "=B"+str(i)+" / F"+str(i)

    ws["K1"] = "Eficiência 4"
    ws["L1"] = "Eficiência 8"
    ws["M1"] = "Eficiência 16"
    ws["N1"] = "Eficiência 32"

    for i in range(2, 10):
        ws["K"+str(i)] = "=G"+str(i)+" / C1"
        ws["L"+str(i)] = "=H"+str(i)+" / D1"
        ws["M"+str(i)] = "=I"+str(i)+" / E1"
        ws["N"+str(i)] = "=J"+str(i)+" / F1"

    for row in ws.rows:
        for c in row:
            c.number_format = '0.0000000000'


    workbook.save(filename=name+".xlsx")

serials = [pd.read_csv(path) for path in path_serials]
parallels = [pd.read_csv(path) for path in path_parallels]

serial = pd.concat(serials)
parallel = pd.concat(parallels)

serial_clean = remove_min_max(serial)
serial_clean = remove_min_max(serial_clean)
parallel_clean = remove_min_max(parallel)
parallel_clean = remove_min_max(parallel_clean)

serial_data = serial_clean.drop(["trial", "cores"], axis=1).groupby(["program", "problem_size"]).mean()
parallel_data = parallel_clean.drop("trial", axis=1).groupby(["program", "cores", "problem_size"]).mean()

def add_col(target, data, name):
    data.columns = [name]
    target = pd.concat([target, data[name]], axis=1)
    return target

without_transpose = serial_data.loc["serial_unit"].reset_index()
without_transpose.columns = ["Tamanho do Problema", "Serial"]
without_transpose = without_transpose.set_index("Tamanho do Problema")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_unit"].loc[4], "4")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_unit"].loc[8], "8")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_unit"].loc[16], "16")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_unit"].loc[32], "32")

save_with_speedup_efic(without_transpose, "parallel_unit")

without_transpose = serial_data.loc["serial_steps"].reset_index()
without_transpose.columns = ["Tamanho do Problema", "Serial"]
without_transpose = without_transpose.set_index("Tamanho do Problema")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_steps"].loc[4], "4")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_steps"].loc[8], "8")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_steps"].loc[16], "16")
without_transpose = add_col(without_transpose, parallel_data.loc["parallel_steps"].loc[32], "32")

save_with_speedup_efic(without_transpose, "parallel_steps")


